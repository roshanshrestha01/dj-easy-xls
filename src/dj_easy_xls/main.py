import re
from django.db import models
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.writer.excel import save_virtual_workbook
import xlrd


def change_format(instance, value):
    field = instance.__class__._meta.get_field(value)
    if isinstance(field, models.ForeignKey):
        return str(getattr(instance, value) or '')
    if isinstance(field, models.ManyToManyField):
        query = list(map(lambda x: x.__str__(), getattr(instance, value).all()))
        text = ', '.join(query)
        return text
    elif isinstance(field, models.DateField):
        if getattr(instance, value) is None:
            return ''
        return getattr(instance, value).strftime('%Y-%m-%d')
    elif isinstance(field, models.BooleanField):
        if getattr(instance, value):
            return 'True'
        return 'False'
    return getattr(instance, value)


class OpenpyxlExport(object):
    def __init__(self, filename='Untitled', row_index=1):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.row_idx = row_index
        self.filename = filename

    def insert_row(self, args):
        for i, value in enumerate(args):
            cell = self.ws.cell(row=self.row_idx, column=i + 1)
            cell.value = value
        self.row_idx += 1

    def set_column_width(self, column, width):
        self.ws.column_dimensions[column].width = width

    def merge_add(self, cell_range, height=None, align='center', font_size=None, bold=False, value='Openpyxl', index=True):
        self.ws.merge_cells(cell_range)
        cell = self.ws.cell(cell_range.split(':')[0])
        cell.value = value
        if height:
            self.ws.row_dimensions[cell.row].height = height
        if font_size:
            cell.font = Font(
                size=font_size,
                bold=bold,
            )
        cell.alignment = Alignment(
            horizontal=align
        )
        if index:
            self.row_idx += 1

    def as_text(self, value):
        if value is None:
            return ""
        return str(value)

    def add_to_cell(self, cell, value):
        self.ws[cell] = value

    def set_width(self):
        for column_cells in self.ws.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            self.ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    def to_camelcase(self, string):
        text = re.sub(r'(?!^)_([a-zA-Z])', lambda m: ' ' + m.group(1).upper(), str(string))
        return text.upper()

    def convert(self, arr):
        return [self.to_camelcase(value) for value in arr]

    def generate(self, arr, convert=False):
        if convert:
            self.insert_row(self.convert(arr))
        else:
            self.insert_row(arr)

    def skip(self, val=1):
        self.row_idx += val

    def response(self):
        response = HttpResponse(save_virtual_workbook(self.wb), content_type='application/vnd.ms-excel')
        file_name = self.filename + '.xlsx'
        response['Content-Disposition'] = 'attachment; filename=' + file_name
        return response
        

class OpenpyxlImport(object):
    def __init__(self, file):
        self.file = file
        if self.file.name.endswith('.xls'):
            self.wb = self.xls_to_xlsx(self.file)
        else:
            self.wb = load_workbook(self.file)
        self.sheets = self.wb.worksheets

    def to_camelcase(self, string):
        text = re.sub(r'(?!^)_([a-zA-Z])', lambda m: ' ' + m.group(1).upper(), str(string))
        return text.upper()

    def to_snake_case(self, string):
        text = re.sub(r'\s', '_', str(string))
        return text.lower()

    def xls_to_xlsx(self, content):
        xls_book = xlrd.open_workbook(file_contents=content.read())
        workbook = openpyxlWorkbook()

        for i in range(0, xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(i)
            sheet = workbook.active if i == 0 else workbook.create_sheet()
            sheet.title = xls_sheet.name

            for row in range(0, xls_sheet.nrows):
                for col in range(0, xls_sheet.ncols):
                    sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)
        return workbook

    def tally_header(self, row, fields):
        return [cell.value for cell in row] == [self.to_camelcase(value) for value in fields]

    def row_to_dict(self, row):
        dct = {}
        for cell in row:
            dct[self.to_snake_case(self.get_first_sheet()[cell.column + '1'].value)] = cell.value
        return dct

    def get_sheets(self):
        return self.sheets

    def get_first_sheet(self):
        return self.sheets[0]

    def get_sheet_rows(self):
        return tuple(self.get_first_sheet().iter_rows())
        